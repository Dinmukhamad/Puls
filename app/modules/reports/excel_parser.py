"""Парсинг Excel-файлов Monthly Report и Report (ТЗ Этап 6, excel_parser.py).

Перенос из services/period_reports.py БЕЗ изменения логики. Здесь критичная
для §16 функция parse_scores: несколько оценок качества в одной ячейке через
запятую разбираются как ОТДЕЛЬНЫЕ оценки, а не как одна строка.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date

import openpyxl

from app.core.datetime_utils import business_today

# Заголовки-агрегаты, которые нельзя путать с датами
_AGGREGATE_HEADERS = {
    "итого", "итого часов", "итого баллов", "итог", "итог часов",
    "средний балл", "кол-во", "план", "всего (ч)", "всего", "отн.",
    "квз", "база часов", "тех. сбои (ч)", "тренинги (ч)",
    "офлайн активность (ч)", "вып нормы (%)", "выработка",
    "ставка", "норма часов (ч)", "оператор", "фио",
}

_DATE_RE = re.compile(r"^(\d{1,2})[.\-/](\d{1,2})(?:[.\-/](\d{2,4}))?$")

# Служебные строки — никогда не считаются операторами
_SERVICE_ROWS = {
    "итого", "другое", "корп такси", "легенда причин", "не выход",
    "опоздание", "причина", "прокси карта", "штраф", "штрафы",
    "комментарий", "итого часов", "всего", "план", "примечание",
    "без причины", "технический сбой", "выходной", "отпуск",
    "больничный", "корпоративное такси",
}


def normalize_name(name: str | None) -> str:
    """ФИО -> нормализованный ключ для сопоставления между файлами и сайтом."""
    if not name:
        return ""
    s = str(name).strip()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("ё", "е").replace("Ё", "Е")
    return s.lower()


def is_service_row(name: str | None) -> bool:
    """True если строка — служебная (итого/причина/опоздание...), не оператор."""
    norm = normalize_name(name)
    if not norm:
        return True
    return norm in _SERVICE_ROWS


def _parse_header_date(value, year: int) -> date | None:
    """Парсит заголовок колонки вида '15.06' в date(year, 6, 15). None если не дата."""
    if value is None:
        return None
    if isinstance(value, date):
        return value
    s = str(value).strip().lower()
    if s in _AGGREGATE_HEADERS:
        return None
    m = _DATE_RE.match(s)
    if not m:
        return None
    day, month = int(m.group(1)), int(m.group(2))
    yr = int(m.group(3)) if m.group(3) else year
    if yr < 100:
        yr += 2000
    try:
        return date(yr, month, day)
    except ValueError:
        return None


def parse_scores(cell_value) -> list[float]:
    """'100, 90, 100' -> [100.0, 90.0, 100.0]. Пустая ячейка -> []."""
    if cell_value is None:
        return []
    s = str(cell_value).strip()
    if not s:
        return []
    out = []
    for part in s.split(","):
        part = part.strip()
        if not part:
            continue
        try:
            out.append(float(part))
        except ValueError:
            continue
    return out


@dataclass
class QualityResult:
    scores: list[float] = field(default_factory=list)
    display_name: str = ""

    @property
    def avg(self) -> float:
        return round(sum(self.scores) / len(self.scores), 2) if self.scores else 0.0

    @property
    def count(self) -> int:
        return len(self.scores)


def parse_monthly_report(
    file_bytes: bytes,
    period_start: date,
    period_end: date,
    default_year: int | None = None,
) -> dict[str, QualityResult]:
    """
    Парсит Monthly Report — несколько листов, на каждом несколько таблиц
    (блоки "ФИО + даты"). Служебные строки отфильтровываются сразу.

    Возвращает {normalized_name: QualityResult}.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    year = default_year or period_start.year
    results: dict[str, QualityResult] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        i = 0
        while i < len(rows):
            row = rows[i]
            if not row or not row[0]:
                i += 1
                continue
            first_cell = str(row[0]).strip().lower()
            if first_cell == "фио":
                header = row
                date_cols: list[tuple[int, date]] = []
                for col_idx, h in enumerate(header):
                    d = _parse_header_date(h, year)
                    if d:
                        date_cols.append((col_idx, d))
                i += 1
                while i < len(rows):
                    data_row = rows[i]
                    if not data_row or not data_row[0]:
                        i += 1
                        continue
                    cell0 = str(data_row[0]).strip()
                    if cell0.lower() == "фио" or "оценки" in cell0.lower():
                        break
                    if is_service_row(cell0):
                        i += 1
                        continue
                    name_key = normalize_name(cell0)
                    if name_key not in results:
                        results[name_key] = QualityResult(display_name=cell0)
                    qr = results[name_key]
                    for col_idx, d in date_cols:
                        if period_start <= d <= period_end and col_idx < len(data_row):
                            qr.scores.extend(parse_scores(data_row[col_idx]))
                    i += 1
                continue
            i += 1

    wb.close()
    return results


def _parse_simple_sheet(
    ws,
    period_start: date,
    period_end: date,
    year: int,
    name_col: int = 0,
) -> dict[str, tuple[str, float]]:
    """
    Общий парсер для листов вида: первая колонка — ФИО, остальные — даты,
    последние колонки — агрегаты. Возвращает {norm_name: (display_name, sum)}.
    Служебные строки отфильтровываются.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = rows[0]
    date_cols: list[tuple[int, date]] = []
    for col_idx, h in enumerate(header):
        if col_idx == name_col:
            continue
        d = _parse_header_date(h, year)
        if d:
            date_cols.append((col_idx, d))

    out: dict[str, tuple[str, float]] = {}
    for row in rows[1:]:
        if not row or not row[name_col]:
            continue
        raw_name = str(row[name_col]).strip()
        if is_service_row(raw_name):
            continue
        name_key = normalize_name(raw_name)
        total = 0.0
        for col_idx, d in date_cols:
            if period_start <= d <= period_end and col_idx < len(row):
                v = row[col_idx]
                if isinstance(v, (int, float)):
                    total += float(v)
                elif isinstance(v, str) and v.strip():
                    # Некоторые листы (например "Штрафы") хранят числа как текст
                    cleaned = v.strip().replace(",", ".").replace(" ", "")
                    try:
                        total += float(cleaned)
                    except ValueError:
                        pass
        prev = out.get(name_key, (raw_name, 0.0))
        out[name_key] = (raw_name, prev[1] + total)
    return out


REQUIRED_REPORT_SHEETS = [
    "Отработанные часы", "Звонки", "Эффективность",
    "Штрафы", "Тренинги", "Тех. сбои", "Офлайн активность",
]


def parse_report_file(
    file_bytes: bytes,
    period_start: date,
    period_end: date,
    default_year: int | None = None,
) -> dict[str, dict[str, tuple[str, float]]]:
    """
    Парсит Report — возвращает { sheet_name: { norm_name: (display_name, sum) } }.
    Бросает ValueError если обязательный лист отсутствует.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    year = default_year or period_start.year

    missing = [s for s in REQUIRED_REPORT_SHEETS if s not in wb.sheetnames]
    if missing:
        wb.close()
        raise ValueError(f"В файле Report отсутствуют листы: {', '.join(missing)}")

    out: dict[str, dict[str, tuple[str, float]]] = {}
    for sheet in REQUIRED_REPORT_SHEETS:
        ws = wb[sheet]
        out[sheet] = _parse_simple_sheet(ws, period_start, period_end, year)

    wb.close()
    return out


# ═══════════════════════════════════════════════════════════════════
# ПОСУТОЧНЫЙ ПАРСИНГ — для сохранения в OperatorDailyMetric
# ═══════════════════════════════════════════════════════════════════
#
# В отличие от parse_monthly_report/parse_report_file (которые сразу
# суммируют значения внутри заданного диапазона), эти функции проходят
# ПО ВСЕМ датам, найденным в файле, и возвращают значения раздельно по
# каждому дню — без диапазона и без агрегации. Используются один раз
# при загрузке файлов (POST /period-report/upload), результат пишется
# в БД построчно на (operator, date).
#
# ВАЖНО: КВЗ и эффективность — производные показатели (звонки/база_часов,
# часы_в_звонке/база_часов × 100), их нельзя честно посчитать на уровне
# одного дня и затем складывать — иначе среднее по дням исказит результат
# относительно правильного "сумма звонков за период / сумма базы часов
# за период". Поэтому в OperatorDailyMetric.kvz/efficiency хранятся
# дневные значения только для информации (heatmap и т.п.), а при агрегации
# произвольного диапазона эти поля ВСЕГДА пересчитываются заново из сумм
# (см. aggregate_daily_metrics).



def parse_monthly_report_daily(
    file_bytes: bytes,
    default_year: int | None = None,
) -> dict[tuple[str, date], dict[str, object]]:
    """
    Возвращает { (name_key, date): {"display_name": str, "scores": [float, ...]} }
    по ВСЕМ датам, найденным в файле (без ограничения диапазоном).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    out: dict[tuple[str, date], dict[str, object]] = {}

    # Определяем "опорный" год по первой найденной дате в файле (если возможно),
    # иначе используем текущий год — это резервный случай для совсем пустых файлов
    year_guess = default_year

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        i = 0
        while i < len(rows):
            row = rows[i]
            if not row or not row[0]:
                i += 1
                continue
            first_cell = str(row[0]).strip().lower()
            if first_cell == "фио":
                header = row
                if year_guess is None:
                    year_guess = business_today().year
                date_cols: list[tuple[int, date]] = []
                for col_idx, h in enumerate(header):
                    d = _parse_header_date(h, year_guess)
                    if d:
                        date_cols.append((col_idx, d))
                i += 1
                while i < len(rows):
                    data_row = rows[i]
                    if not data_row or not data_row[0]:
                        i += 1
                        continue
                    cell0 = str(data_row[0]).strip()
                    if cell0.lower() == "фио" or "оценки" in cell0.lower():
                        break
                    if is_service_row(cell0):
                        i += 1
                        continue
                    name_key = normalize_name(cell0)
                    for col_idx, d in date_cols:
                        if col_idx >= len(data_row):
                            continue
                        scores = parse_scores(data_row[col_idx])
                        if not scores:
                            continue
                        key = (name_key, d)
                        if key not in out:
                            out[key] = {"display_name": cell0, "scores": []}
                        out[key]["scores"].extend(scores)
                    i += 1
                continue
            i += 1

    wb.close()
    return out


def _parse_simple_sheet_daily(ws, year: int, name_col: int = 0) -> dict[tuple[str, date], tuple[str, float]]:
    """Версия _parse_simple_sheet без диапазона — все даты, по отдельности."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = rows[0]
    date_cols: list[tuple[int, date]] = []
    for col_idx, h in enumerate(header):
        if col_idx == name_col:
            continue
        d = _parse_header_date(h, year)
        if d:
            date_cols.append((col_idx, d))

    out: dict[tuple[str, date], tuple[str, float]] = {}
    for row in rows[1:]:
        if not row or not row[name_col]:
            continue
        raw_name = str(row[name_col]).strip()
        if is_service_row(raw_name):
            continue
        name_key = normalize_name(raw_name)
        for col_idx, d in date_cols:
            if col_idx >= len(row):
                continue
            v = row[col_idx]
            value = 0.0
            if isinstance(v, (int, float)):
                value = float(v)
            elif isinstance(v, str) and v.strip():
                cleaned = v.strip().replace(",", ".").replace(" ", "")
                try:
                    value = float(cleaned)
                except ValueError:
                    continue
            else:
                continue
            out[(name_key, d)] = (raw_name, value)
    return out


def parse_report_file_daily(
    file_bytes: bytes,
    default_year: int | None = None,
) -> dict[str, dict[tuple[str, date], tuple[str, float]]]:
    """
    Возвращает { sheet_name: { (name_key, date): (display_name, value) } }
    по ВСЕМ датам в файле. Бросает ValueError если обязательный лист отсутствует.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    year = default_year or business_today().year

    missing = [s for s in REQUIRED_REPORT_SHEETS if s not in wb.sheetnames]
    if missing:
        wb.close()
        raise ValueError(f"В файле Report отсутствуют листы: {', '.join(missing)}")

    out: dict[str, dict[tuple[str, date], tuple[str, float]]] = {}
    for sheet in REQUIRED_REPORT_SHEETS:
        ws = wb[sheet]
        out[sheet] = _parse_simple_sheet_daily(ws, year)

    wb.close()
    return out
