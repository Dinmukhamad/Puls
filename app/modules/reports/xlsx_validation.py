from __future__ import annotations

from io import BytesIO
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

MAX_ARCHIVE_ENTRIES = 250
MAX_UNCOMPRESSED_BYTES = 120 * 1024 * 1024
MAX_WORKSHEETS = 30
MAX_ROWS_PER_SHEET = 250_000
MAX_COLUMNS_PER_SHEET = 250
MAX_CELLS = 2_000_000


def validate_xlsx_archive(content: bytes, label: str) -> None:
    """Validate the XLSX container before openpyxl parses any workbook XML."""
    if not content.startswith(b"PK"):
        raise ValueError(f"{label}: файл не является XLSX-архивом")

    try:
        with ZipFile(BytesIO(content)) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_ARCHIVE_ENTRIES:
                raise ValueError(f"{label}: слишком много файлов внутри XLSX")
            total_uncompressed = sum(entry.file_size for entry in entries)
            if total_uncompressed > MAX_UNCOMPRESSED_BYTES:
                raise ValueError(f"{label}: распакованный XLSX превышает безопасный размер")
            if not {"[Content_Types].xml", "xl/workbook.xml"}.issubset(
                {entry.filename for entry in entries}
            ):
                raise ValueError(f"{label}: повреждена структура XLSX")
    except BadZipFile as exc:
        raise ValueError(f"{label}: повреждённый XLSX-архив") from exc

    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise ValueError(f"{label}: книгу Excel невозможно открыть") from exc

    try:
        if not workbook.worksheets:
            raise ValueError(f"{label}: в книге нет листов")
        if len(workbook.worksheets) > MAX_WORKSHEETS:
            raise ValueError(f"{label}: слишком много листов")
        cells = 0
        for worksheet in workbook.worksheets:
            if worksheet.max_row > MAX_ROWS_PER_SHEET:
                raise ValueError(f"{label}: лист «{worksheet.title}» содержит слишком много строк")
            if worksheet.max_column > MAX_COLUMNS_PER_SHEET:
                raise ValueError(f"{label}: лист «{worksheet.title}» содержит слишком много столбцов")
            cells += worksheet.max_row * worksheet.max_column
            if cells > MAX_CELLS:
                raise ValueError(f"{label}: книга содержит слишком много ячеек")
    finally:
        workbook.close()
