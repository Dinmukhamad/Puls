"""
Сервис расчёта показателей операторов по загруженным Excel-файлам за выбранный период.

Источники:
  - Monthly Report: оценки качества звонков (несколько листов-проверяющих, одна
    таблица операторов на каждом, колонки = даты, ячейки = "100, 90, 100").
  - Report: отдельные листы — "Отработанные часы", "Звонки", "Эффективность",
    "Штрафы", "Тренинги", "Тех. сбои", "Офлайн активность". Колонки = даты,
    последняя колонка(и) = агрегаты ("Итого", "Всего (ч)" и т.п.) — игнорируются.

Ключевое правило: в расчёт сводных показателей (summary) попадают ТОЛЬКО
операторы, которые одновременно есть на сайте (siteOperators) и в файле.
Служебные строки (итого, причина, опоздание...) отфильтровываются ещё на
этапе парсинга и никогда не считаются операторами.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date

import openpyxl

# Заголовки-агрегаты, которые нельзя путать с датами
_AGGREGATE_HEADERS = {
    "итого", "итого часов", "итого баллов", "итог", "итог часов",
    "средний балл", "кол-во", "план", "всего (ч)", "всего", "отн.",
    "квз", "база часов", "тех. сбои (ч)", "тренинги (ч)",
    "офлайн активность (ч)", "вып нормы (%)", "выработка",
    "ставка", "норма часов (ч)", "оператор", "фио",
}

_DATE_RE = re.compile(r"^(\d{1,2})[.\-/](\d{1,2})(?:[.\-/](\d{2,4}))?$")

# Служебные строки — никогда не считаются операторами
_SERVICE_ROWS = {
    "итого", "другое", "корп такси", "легенда причин", "не выход",
    "опоздание", "причина", "прокси карта", "штраф", "штрафы",
    "комментарий", "итого часов", "всего", "план", "примечание",
    "без причины", "технический сбой", "выходной", "отпуск",
    "больничный", "корпоративное такси",
}


def normalize_name(name: str | None) -> str:
    """ФИО -> нормализованный ключ для сопоставления между файлами и сайтом."""
    if not name:
        return ""
    s = str(name).strip()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("ё", "е").replace("Ё", "Е")
    return s.lower()


def is_service_row(name: str | None) -> bool:
    """True если строка — служебная (итого/причина/опоздание...), не оператор."""
    norm = normalize_name(name)
    if not norm:
        return True
    return norm in _SERVICE_ROWS


def _parse_header_date(value, year: int) -> date | None:
    """Парсит заголовок колонки вида '15.06' в date(year, 6, 15). None если не дата."""
    if value is None:
        return None
    if isinstance(value, date):
        return value
    s = str(value).strip().lower()
    if s in _AGGREGATE_HEADERS:
        return None
    m = _DATE_RE.match(s)
    if not m:
        return None
    day, month = int(m.group(1)), int(m.group(2))
    yr = int(m.group(3)) if m.group(3) else year
    if yr < 100:
        yr += 2000
    try:
        return date(yr, month, day)
    except ValueError:
        return None


def parse_scores(cell_value) -> list[float]:
    """'100, 90, 100' -> [100.0, 90.0, 100.0]. Пустая ячейка -> []."""
    if cell_value is None:
        return []
    s = str(cell_value).strip()
    if not s:
        return []
    out = []
    for part in s.split(","):
        part = part.strip()
        if not part:
            continue
        try:
            out.append(float(part))
        except ValueError:
            continue
    return out


@dataclass
class QualityResult:
    scores: list[float] = field(default_factory=list)
    display_name: str = ""

    @property
    def avg(self) -> float:
        return round(sum(self.scores) / len(self.scores), 2) if self.scores else 0.0

    @property
    def count(self) -> int:
        return len(self.scores)


def parse_monthly_report(
    file_bytes: bytes,
    period_start: date,
    period_end: date,
    default_year: int | None = None,
) -> dict[str, QualityResult]:
    """
    Парсит Monthly Report — несколько листов, на каждом несколько таблиц
    (блоки "ФИО + даты"). Служебные строки отфильтровываются сразу.

    Возвращает {normalized_name: QualityResult}.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    year = default_year or period_start.year
    results: dict[str, QualityResult] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        i = 0
        while i < len(rows):
            row = rows[i]
            if not row or not row[0]:
                i += 1
                continue
            first_cell = str(row[0]).strip().lower()
            if first_cell == "фио":
                header = row
                date_cols: list[tuple[int, date]] = []
                for col_idx, h in enumerate(header):
                    d = _parse_header_date(h, year)
                    if d:
                        date_cols.append((col_idx, d))
                i += 1
                while i < len(rows):
                    data_row = rows[i]
                    if not data_row or not data_row[0]:
                        i += 1
                        continue
                    cell0 = str(data_row[0]).strip()
                    if cell0.lower() == "фио" or "оценки" in cell0.lower():
                        break
                    if is_service_row(cell0):
                        i += 1
                        continue
                    name_key = normalize_name(cell0)
                    if name_key not in results:
                        results[name_key] = QualityResult(display_name=cell0)
                    qr = results[name_key]
                    for col_idx, d in date_cols:
                        if period_start <= d <= period_end and col_idx < len(data_row):
                            qr.scores.extend(parse_scores(data_row[col_idx]))
                    i += 1
                continue
            i += 1

    wb.close()
    return results


def _parse_simple_sheet(
    ws,
    period_start: date,
    period_end: date,
    year: int,
    name_col: int = 0,
) -> dict[str, tuple[str, float]]:
    """
    Общий парсер для листов вида: первая колонка — ФИО, остальные — даты,
    последние колонки — агрегаты. Возвращает {norm_name: (display_name, sum)}.
    Служебные строки отфильтровываются.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = rows[0]
    date_cols: list[tuple[int, date]] = []
    for col_idx, h in enumerate(header):
        if col_idx == name_col:
            continue
        d = _parse_header_date(h, year)
        if d:
            date_cols.append((col_idx, d))

    out: dict[str, tuple[str, float]] = {}
    for row in rows[1:]:
        if not row or not row[name_col]:
            continue
        raw_name = str(row[name_col]).strip()
        if is_service_row(raw_name):
            continue
        name_key = normalize_name(raw_name)
        total = 0.0
        for col_idx, d in date_cols:
            if period_start <= d <= period_end and col_idx < len(row):
                v = row[col_idx]
                if isinstance(v, (int, float)):
                    total += float(v)
                elif isinstance(v, str) and v.strip():
                    # Некоторые листы (например "Штрафы") хранят числа как текст
                    cleaned = v.strip().replace(",", ".").replace(" ", "")
                    try:
                        total += float(cleaned)
                    except ValueError:
                        pass
        prev = out.get(name_key, (raw_name, 0.0))
        out[name_key] = (raw_name, prev[1] + total)
    return out


REQUIRED_REPORT_SHEETS = [
    "Отработанные часы", "Звонки", "Эффективность",
    "Штрафы", "Тренинги", "Тех. сбои", "Офлайн активность",
]


def parse_report_file(
    file_bytes: bytes,
    period_start: date,
    period_end: date,
    default_year: int | None = None,
) -> dict[str, dict[str, tuple[str, float]]]:
    """
    Парсит Report — возвращает { sheet_name: { norm_name: (display_name, sum) } }.
    Бросает ValueError если обязательный лист отсутствует.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    year = default_year or period_start.year

    missing = [s for s in REQUIRED_REPORT_SHEETS if s not in wb.sheetnames]
    if missing:
        wb.close()
        raise ValueError(f"В файле Report отсутствуют листы: {', '.join(missing)}")

    out: dict[str, dict[str, tuple[str, float]]] = {}
    for sheet in REQUIRED_REPORT_SHEETS:
        ws = wb[sheet]
        out[sheet] = _parse_simple_sheet(ws, period_start, period_end, year)

    wb.close()
    return out


@dataclass
class OperatorPeriodMetrics:
    full_name: str
    name_key: str = ""
    quality_avg: float = 0.0
    quality_calls_count: int = 0
    quality_scores: list[float] = field(default_factory=list)
    total_hours: float = 0.0
    base_hours: float = 0.0
    tech_issue_hours: float = 0.0
    training_hours: float = 0.0
    offline_activity_hours: float = 0.0
    calls_total: float = 0.0
    kvz: float = 0.0
    call_time_hours: float = 0.0
    efficiency_percent: float = 0.0
    penalty_sum: float = 0.0
    penalty_minutes: float = 0.0
    penalty_points: float = 0.0
    final_points: float = 0.0
    has_any_period_data: bool = False
    warnings: list[str] = field(default_factory=list)
    # Поля нормы часов (заполняются отдельно через work_norms сервис)
    rate: float | None = None
    individual_norm_hours: float = 0.0
    norm_completion_percent: float = 0.0
    hours_points: float = 0.0
    overtime_hours: float = 0.0
    overtime_percent: float = 0.0
    norm_warnings: list[str] = field(default_factory=list)


PENALTY_RUB_PER_MINUTE = 50.0
PENALTY_POINTS_PER_MINUTE = 5.0


def compute_operator_metrics(
    name_key: str,
    display_name: str,
    quality: QualityResult | None,
    report_data: dict[str, dict[str, tuple[str, float]]],
) -> OperatorPeriodMetrics:
    m = OperatorPeriodMetrics(full_name=display_name, name_key=name_key)

    # Качество
    if quality and quality.scores:
        m.quality_avg = quality.avg
        m.quality_calls_count = quality.count
        m.quality_scores = list(quality.scores)
    else:
        m.warnings.append("Нет оценок качества за выбранный период")

    def sheet_val(sheet: str) -> float:
        entry = report_data.get(sheet, {}).get(name_key)
        return entry[1] if entry else 0.0

    m.total_hours = round(sheet_val("Отработанные часы"), 2)
    m.tech_issue_hours = round(sheet_val("Тех. сбои"), 2)
    m.training_hours = round(sheet_val("Тренинги"), 2)
    m.offline_activity_hours = round(sheet_val("Офлайн активность"), 2)

    base = m.total_hours - m.tech_issue_hours - m.training_hours - m.offline_activity_hours
    if base < 0:
        m.warnings.append("База часов получилась отрицательной. Проверьте тренинги, техсбои и офлайн-активность.")
        base = 0.0
    m.base_hours = round(base, 2)

    m.calls_total = round(sheet_val("Звонки"), 2)
    if m.base_hours > 0:
        m.kvz = round(m.calls_total / m.base_hours, 2)
    else:
        m.kvz = 0.0
        m.warnings.append("Нет базы часов за выбранный период")

    m.call_time_hours = round(sheet_val("Эффективность"), 2)
    if m.base_hours > 0:
        m.efficiency_percent = round(m.call_time_hours / m.base_hours * 100, 2)
    else:
        m.efficiency_percent = 0.0
        if "Нет базы часов за выбранный период" not in m.warnings:
            m.warnings.append("Нет базы часов для расчёта эффективности")

    m.penalty_sum = round(sheet_val("Штрафы"), 2)
    m.penalty_minutes = round(m.penalty_sum / PENALTY_RUB_PER_MINUTE, 2) if m.penalty_sum else 0.0
    m.penalty_points = round(m.penalty_minutes * PENALTY_POINTS_PER_MINUTE, 2)

    # Базовые итоговые баллы БЕЗ часов — norm-aware финальный расчёт
    # делается в роутере после enrich_with_norm(). Сохраняем 0 за часы
    # чтобы роутер мог корректно подставить hours_points.
    m.final_points = round(
        m.quality_avg + m.kvz + m.total_hours + m.efficiency_percent - m.penalty_points,
        2,
    )
    # Пометим что hours_points ещё не заполнен — роутер подставит позже

    m.has_any_period_data = any([
        m.quality_calls_count > 0,
        m.total_hours > 0,
        m.calls_total > 0,
        m.base_hours > 0,
        m.penalty_sum > 0,
    ])

    return m


@dataclass
class PeriodCalculationResult:
    operators: list[OperatorPeriodMetrics]              # только matched, с данными
    warnings_site_only: list[str]                        # есть на сайте, нет в файле
    warnings_file_only: list[str]                        # есть в файле, нет на сайте
    warnings_no_quality: list[str]                       # нет оценок качества
    warnings_no_base_hours: list[str]                    # нет базы часов
    ignored_service_rows: list[str]                      # игнорированные служебные строки
    summary: dict[str, float | None]                  # сводные показатели


def calculate_period_report(
    monthly_report_bytes: bytes,
    report_bytes: bytes,
    period_start: date,
    period_end: date,
    site_operator_names: list[str],
) -> PeriodCalculationResult:
    """
    site_operator_names — список full_name операторов из БД сайта.
    Используется для построения matched-выборки: в расчёт идут только те,
    кто есть и на сайте, и в файле.
    """
    if period_start > period_end:
        raise ValueError("Дата начала не может быть позже даты окончания")

    quality_map = parse_monthly_report(monthly_report_bytes, period_start, period_end)
    report_data = parse_report_file(report_bytes, period_start, period_end)

    # Множество имён операторов сайта (нормализованных)
    site_keys = {normalize_name(n): n for n in site_operator_names if n and not is_service_row(n)}

    # Все ключи из файлов (только реальные, не служебные — уже отфильтровано на парсинге)
    file_keys: dict[str, str] = {}
    for key, qr in quality_map.items():
        file_keys.setdefault(key, qr.display_name)
    for sheet_data in report_data.values():
        for key, (disp, _val) in sheet_data.items():
            file_keys.setdefault(key, disp)

    matched_keys = set(site_keys.keys()) & set(file_keys.keys())
    site_only_keys = set(site_keys.keys()) - set(file_keys.keys())
    file_only_keys = set(file_keys.keys()) - set(site_keys.keys())

    operators: list[OperatorPeriodMetrics] = []
    warnings_no_quality: list[str] = []
    warnings_no_base_hours: list[str] = []

    for key in sorted(matched_keys):
        display = site_keys.get(key) or file_keys.get(key) or key
        q = quality_map.get(key)
        metrics = compute_operator_metrics(key, display, q, report_data)

        if not metrics.has_any_period_data:
            # Matched, но реально нет никаких данных за период — не включаем в расчёт
            continue

        operators.append(metrics)
        if "Нет оценок качества за выбранный период" in metrics.warnings:
            warnings_no_quality.append(display)
        if "Нет базы часов за выбранный период" in metrics.warnings:
            warnings_no_base_hours.append(display)

    operators.sort(key=lambda m: m.final_points, reverse=True)

    warnings_site_only = sorted(site_keys[k] for k in site_only_keys)
    warnings_file_only = sorted(file_keys[k] for k in file_only_keys)

    # Сводные показатели — считаем ТОЛЬКО по matched + has_any_period_data
    included = operators  # уже отфильтрованы выше

    all_quality_scores: list[float] = []
    for op in included:
        all_quality_scores.extend(op.quality_scores)
    avg_quality = round(sum(all_quality_scores) / len(all_quality_scores), 2) if all_quality_scores else None

    total_calls = sum(op.calls_total for op in included)
    total_base_hours = sum(op.base_hours for op in included if op.base_hours > 0)
    total_call_time = sum(op.call_time_hours for op in included if op.base_hours > 0)
    total_penalty_sum = sum(op.penalty_sum for op in included)

    avg_kvz = round(total_calls / total_base_hours, 2) if total_base_hours > 0 else None
    avg_efficiency = round(total_call_time / total_base_hours * 100, 2) if total_base_hours > 0 else None
    penalty_minutes_total = round(total_penalty_sum / PENALTY_RUB_PER_MINUTE, 2) if total_penalty_sum else 0.0

    summary = {
        "operators_count": len(included),
        "site_total_count": len(site_keys),
        "matched_count": len(matched_keys),
        "site_only_count": len(site_only_keys),
        "file_only_count": len(file_only_keys),
        "avg_quality": avg_quality,
        "total_calls": round(total_calls, 2),
        "avg_kvz": avg_kvz,
        "avg_efficiency": avg_efficiency,
        "penalty_minutes_total": penalty_minutes_total,
    }

    return PeriodCalculationResult(
        operators=operators,
        warnings_site_only=warnings_site_only,
        warnings_file_only=warnings_file_only,
        warnings_no_quality=warnings_no_quality,
        warnings_no_base_hours=warnings_no_base_hours,
        ignored_service_rows=sorted(_SERVICE_ROWS),
        summary=summary,
    )


# ═══════════════════════════════════════════════════════════════════
# ПОСУТОЧНЫЙ ПАРСИНГ — для сохранения в OperatorDailyMetric
# ═══════════════════════════════════════════════════════════════════
#
# В отличие от parse_monthly_report/parse_report_file (которые сразу
# суммируют значения внутри заданного диапазона), эти функции проходят
# ПО ВСЕМ датам, найденным в файле, и возвращают значения раздельно по
# каждому дню — без диапазона и без агрегации. Используются один раз
# при загрузке файлов (POST /period-report/upload), результат пишется
# в БД построчно на (operator, date).
#
# ВАЖНО: КВЗ и эффективность — производные показатели (звонки/база_часов,
# часы_в_звонке/база_часов × 100), их нельзя честно посчитать на уровне
# одного дня и затем складывать — иначе среднее по дням исказит результат
# относительно правильного "сумма звонков за период / сумма базы часов
# за период". Поэтому в OperatorDailyMetric.kvz/efficiency хранятся
# дневные значения только для информации (heatmap и т.п.), а при агрегации
# произвольного диапазона эти поля ВСЕГДА пересчитываются заново из сумм
# (см. aggregate_daily_metrics).



def parse_monthly_report_daily(
    file_bytes: bytes,
    default_year: int | None = None,
) -> dict[tuple[str, date], dict[str, object]]:
    """
    Возвращает { (name_key, date): {"display_name": str, "scores": [float, ...]} }
    по ВСЕМ датам, найденным в файле (без ограничения диапазоном).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    out: dict[tuple[str, date], dict[str, object]] = {}

    # Определяем "опорный" год по первой найденной дате в файле (если возможно),
    # иначе используем текущий год — это резервный случай для совсем пустых файлов
    year_guess = default_year

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        i = 0
        while i < len(rows):
            row = rows[i]
            if not row or not row[0]:
                i += 1
                continue
            first_cell = str(row[0]).strip().lower()
            if first_cell == "фио":
                header = row
                if year_guess is None:
                    year_guess = date.today().year
                date_cols: list[tuple[int, date]] = []
                for col_idx, h in enumerate(header):
                    d = _parse_header_date(h, year_guess)
                    if d:
                        date_cols.append((col_idx, d))
                i += 1
                while i < len(rows):
                    data_row = rows[i]
                    if not data_row or not data_row[0]:
                        i += 1
                        continue
                    cell0 = str(data_row[0]).strip()
                    if cell0.lower() == "фио" or "оценки" in cell0.lower():
                        break
                    if is_service_row(cell0):
                        i += 1
                        continue
                    name_key = normalize_name(cell0)
                    for col_idx, d in date_cols:
                        if col_idx >= len(data_row):
                            continue
                        scores = parse_scores(data_row[col_idx])
                        if not scores:
                            continue
                        key = (name_key, d)
                        if key not in out:
                            out[key] = {"display_name": cell0, "scores": []}
                        out[key]["scores"].extend(scores)
                    i += 1
                continue
            i += 1

    wb.close()
    return out


def _parse_simple_sheet_daily(ws, year: int, name_col: int = 0) -> dict[tuple[str, date], tuple[str, float]]:
    """Версия _parse_simple_sheet без диапазона — все даты, по отдельности."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = rows[0]
    date_cols: list[tuple[int, date]] = []
    for col_idx, h in enumerate(header):
        if col_idx == name_col:
            continue
        d = _parse_header_date(h, year)
        if d:
            date_cols.append((col_idx, d))

    out: dict[tuple[str, date], tuple[str, float]] = {}
    for row in rows[1:]:
        if not row or not row[name_col]:
            continue
        raw_name = str(row[name_col]).strip()
        if is_service_row(raw_name):
            continue
        name_key = normalize_name(raw_name)
        for col_idx, d in date_cols:
            if col_idx >= len(row):
                continue
            v = row[col_idx]
            value = 0.0
            if isinstance(v, (int, float)):
                value = float(v)
            elif isinstance(v, str) and v.strip():
                cleaned = v.strip().replace(",", ".").replace(" ", "")
                try:
                    value = float(cleaned)
                except ValueError:
                    continue
            else:
                continue
            out[(name_key, d)] = (raw_name, value)
    return out


def parse_report_file_daily(
    file_bytes: bytes,
    default_year: int | None = None,
) -> dict[str, dict[tuple[str, date], tuple[str, float]]]:
    """
    Возвращает { sheet_name: { (name_key, date): (display_name, value) } }
    по ВСЕМ датам в файле. Бросает ValueError если обязательный лист отсутствует.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    year = default_year or date.today().year

    missing = [s for s in REQUIRED_REPORT_SHEETS if s not in wb.sheetnames]
    if missing:
        wb.close()
        raise ValueError(f"В файле Report отсутствуют листы: {', '.join(missing)}")

    out: dict[str, dict[tuple[str, date], tuple[str, float]]] = {}
    for sheet in REQUIRED_REPORT_SHEETS:
        ws = wb[sheet]
        out[sheet] = _parse_simple_sheet_daily(ws, year)

    wb.close()
    return out


@dataclass
class DailyMetricRow:
    name_key: str
    display_name: str
    metric_date: date
    calls_count: float = 0.0
    quality_scores: list[float] = field(default_factory=list)
    worked_hours: float = 0.0
    tech_issue_hours: float = 0.0
    training_hours: float = 0.0
    offline_activity_hours: float = 0.0
    call_time_hours: float = 0.0  # лист "Эффективность" — часы в звонке за день
    penalty_sum: float = 0.0


def build_daily_metric_rows(
    monthly_bytes: bytes,
    report_bytes: bytes,
    default_year: int | None = None,
) -> list[DailyMetricRow]:
    """
    Главная точка входа для посуточного парсинга. Объединяет Monthly Report
    (оценки качества) и Report (часы/звонки/штрафы) в единый список строк
    "оператор × дата", готовых для сохранения в OperatorDailyMetric.

    Вызывается ОДИН раз при загрузке файлов — не при построении аналитики.
    """
    quality_by_day = parse_monthly_report_daily(monthly_bytes, default_year)
    report_by_sheet = parse_report_file_daily(report_bytes, default_year)

    keys = set(quality_by_day.keys())
    for sheet_data in report_by_sheet.values():
        keys |= set(sheet_data.keys())

    display_names: dict[str, str] = {}
    for (name_key, _d), data in quality_by_day.items():
        display_names.setdefault(name_key, data["display_name"])
    for sheet_data in report_by_sheet.values():
        for (name_key, _d), (disp, _val) in sheet_data.items():
            display_names.setdefault(name_key, disp)

    rows: list[DailyMetricRow] = []
    for name_key, metric_date in sorted(keys):
        q = quality_by_day.get((name_key, metric_date))
        row = DailyMetricRow(
            name_key=name_key,
            display_name=display_names.get(name_key, name_key),
            metric_date=metric_date,
            quality_scores=list(q["scores"]) if q else [],
            calls_count=report_by_sheet.get("Звонки", {}).get((name_key, metric_date), (None, 0.0))[1],
            worked_hours=report_by_sheet.get("Отработанные часы", {}).get((name_key, metric_date), (None, 0.0))[1],
            tech_issue_hours=report_by_sheet.get("Тех. сбои", {}).get((name_key, metric_date), (None, 0.0))[1],
            training_hours=report_by_sheet.get("Тренинги", {}).get((name_key, metric_date), (None, 0.0))[1],
            offline_activity_hours=report_by_sheet.get("Офлайн активность", {}).get((name_key, metric_date), (None, 0.0))[1],
            call_time_hours=report_by_sheet.get("Эффективность", {}).get((name_key, metric_date), (None, 0.0))[1],
            penalty_sum=report_by_sheet.get("Штрафы", {}).get((name_key, metric_date), (None, 0.0))[1],
        )
        rows.append(row)

    return rows


def aggregate_daily_rows(daily_rows: list[dict]) -> OperatorPeriodMetrics:
    """
    Агрегирует список словарей-строк OperatorDailyMetric (за произвольный
    диапазон дат, для ОДНОГО оператора) в OperatorPeriodMetrics — те же
    формулы, что в compute_operator_metrics, но источник — БД, не Excel.

    daily_rows: список dict с полями calls_count, quality_sum, quality_count,
    worked_hours, tech_issue_hours, training_hours, offline_activity_hours,
    efficiency (= call_time_hours за день), penalty_sum.
    """
    if not daily_rows:
        return OperatorPeriodMetrics(full_name="", warnings=["Нет данных за выбранный период"])

    quality_sum = sum(r["quality_sum"] for r in daily_rows)
    quality_count = sum(r["quality_count"] for r in daily_rows)

    total_hours = round(sum(r["worked_hours"] for r in daily_rows), 2)
    tech_issue_hours = round(sum(r["tech_issue_hours"] for r in daily_rows), 2)
    training_hours = round(sum(r["training_hours"] for r in daily_rows), 2)
    offline_activity_hours = round(sum(r["offline_activity_hours"] for r in daily_rows), 2)
    calls_total = round(sum(r["calls_count"] for r in daily_rows), 2)
    call_time_hours = round(sum(r["efficiency"] for r in daily_rows), 2)  # "efficiency" в дневной таблице = часы в звонке за день
    penalty_sum = round(sum(r["penalty_sum"] for r in daily_rows), 2)

    m = OperatorPeriodMetrics(full_name="")
    m.quality_calls_count = quality_count
    m.quality_avg = round(quality_sum / quality_count, 2) if quality_count else 0.0
    if quality_count == 0:
        m.warnings.append("Нет оценок качества за выбранный период")

    m.total_hours = total_hours
    m.tech_issue_hours = tech_issue_hours
    m.training_hours = training_hours
    m.offline_activity_hours = offline_activity_hours

    base = total_hours - tech_issue_hours - training_hours - offline_activity_hours
    if base < 0:
        m.warnings.append("База часов получилась отрицательной. Проверьте тренинги, техсбои и офлайн-активность.")
        base = 0.0
    m.base_hours = round(base, 2)

    m.calls_total = calls_total
    if m.base_hours > 0:
        m.kvz = round(calls_total / m.base_hours, 2)
    else:
        m.kvz = 0.0
        m.warnings.append("Нет базы часов за выбранный период")

    m.call_time_hours = call_time_hours
    if m.base_hours > 0:
        m.efficiency_percent = round(call_time_hours / m.base_hours * 100, 2)
    else:
        m.efficiency_percent = 0.0
        if "Нет базы часов за выбранный период" not in m.warnings:
            m.warnings.append("Нет базы часов для расчёта эффективности")

    m.penalty_sum = penalty_sum
    m.penalty_minutes = round(penalty_sum / PENALTY_RUB_PER_MINUTE, 2) if penalty_sum else 0.0
    m.penalty_points = round(m.penalty_minutes * PENALTY_POINTS_PER_MINUTE, 2)

    # Базовые итоговые баллы БЕЗ часов — norm-aware финальный расчёт
    # делается в роутере после enrich_with_norm(). Сохраняем 0 за часы
    # чтобы роутер мог корректно подставить hours_points.
    m.final_points = round(
        m.quality_avg + m.kvz + m.total_hours + m.efficiency_percent - m.penalty_points,
        2,
    )
    # Пометим что hours_points ещё не заполнен — роутер подставит позже

    m.has_any_period_data = any([
        m.quality_calls_count > 0,
        m.total_hours > 0,
        m.calls_total > 0,
        m.base_hours > 0,
        m.penalty_sum > 0,
    ])

    return m
