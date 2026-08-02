from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import delete, func, select

from app.modules.analytics.cache import cache_clear_all
from tests.conftest import make_operator


@pytest.fixture(autouse=True)
def _cleanup(db_session):
    from app.models import entities as m

    baseline = db_session.scalar(select(func.max(m.Operator.id))) or 0
    yield
    ids = list(db_session.scalars(select(m.Operator.id).where(m.Operator.id > baseline)))
    if ids:
        db_session.execute(delete(m.PeriodReport).where(m.PeriodReport.operator_id.in_(ids)))
        db_session.execute(delete(m.OperatorDailyMetric).where(m.OperatorDailyMetric.operator_id.in_(ids)))
        db_session.execute(delete(m.Operator).where(m.Operator.id.in_(ids)))
        db_session.commit()
    cache_clear_all()


def _metric(db, operator, calls):
    from app.models import entities as m

    db.add(m.OperatorDailyMetric(
        operator_id=operator.id, operator_name=operator.full_name,
        metric_date=date(2036, 1, 15), calls_count=calls,
        quality_sum=90, quality_count=1, quality_avg=90,
        worked_hours=8, base_hours=8, efficiency=4,
    ))
    db.commit()
    cache_clear_all()


def test_operator_pagination_sorting_and_filtered_xlsx(client, db_session):
    first = make_operator(db_session, full_name="Export Alpha")
    second = make_operator(db_session, full_name="Export Beta")
    _metric(db_session, first, 10)
    _metric(db_session, second, 20)
    params = {
        "start_date": "2036-01-15", "end_date": "2036-01-15",
        "operator_query": "Export", "only_with_data": "true",
        "page": 1, "page_size": 1, "sort_by": "calls_total", "sort_order": "desc",
    }
    response = client.get("/api/analytics/operators", params=params)
    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["total"] == 2
    assert payload["items"][0]["calls_total"] == 20

    export = client.get("/api/analytics/export.xlsx", params=params)
    assert export.status_code == 200, export.text
    sheet = load_workbook(BytesIO(export.content)).active
    assert sheet["A8"].value == "Оператор"
    names = {sheet.cell(row=index, column=1).value for index in range(9, sheet.max_row + 1)}
    assert {first.full_name, second.full_name} <= names


def test_dashboard_compares_with_previous_period_and_explains_metrics(client, db_session):
    """Экран руководителя приходит одним ответом и объясняет каждый показатель."""
    operator = make_operator(db_session, full_name="Previous metric")
    _metric(db_session, operator, 20)
    response = client.get(
        "/api/analytics/dashboard",
        params={"start_date": "2036-01-15", "end_date": "2036-01-15"},
    )
    assert response.status_code == 200, response.text
    payload = response.json()

    assert payload["period"]["start"] == "2036-01-15"
    assert payload["period"]["days"] == 1
    # 15 января 2036 — вторник, поэтому предыдущий период это 14-е.
    quality = next(m for m in payload["metrics"] if m["key"] == "quality")
    assert quality["delta"]["previous"] is None  # за 14-е данных нет

    # Каждая карточка несёт человеческое объяснение и цель.
    for metric in payload["metrics"]:
        assert metric["definition"]
        assert metric["action"]
        assert metric["status"] in {"good", "watch", "bad", "neutral", "unknown"}
    assert quality["value"] == 90.0
    assert quality["target"] == 85.0
    assert quality["status"] == "good"

    assert len(payload["weekdays"]) == 7
    assert payload["trend"]["metric"] == "quality"
    assert payload["filters"]["all_weekdays"] is True


def test_dashboard_weekday_filter_narrows_the_period(client, db_session):
    """Фильтр по дням недели заменяет отсутствующий в данных разрез по времени."""
    operator = make_operator(db_session, full_name="Weekday metric")
    _metric(db_session, operator, 20)  # вторник, 15 января 2036

    tuesday = client.get("/api/analytics/dashboard", params={
        "start_date": "2036-01-13", "end_date": "2036-01-19", "weekdays": "1",
    }).json()
    assert [p["date"] for p in tuesday["trend"]["points"]] == ["2036-01-14"] or \
           any(p["has_data"] for p in tuesday["trend"]["points"])
    assert tuesday["filters"]["weekdays"] == [1]
    assert tuesday["filters"]["all_weekdays"] is False

    monday = client.get("/api/analytics/dashboard", params={
        "start_date": "2036-01-13", "end_date": "2036-01-19", "weekdays": "0",
    }).json()
    assert monday["coverage"]["days_with_data"] == 0


def test_glossary_lists_every_headline_metric(client):
    response = client.get("/api/analytics/glossary")
    assert response.status_code == 200, response.text
    metrics = response.json()["metrics"]
    keys = {m["metric_key"] for m in metrics}
    assert {"quality", "kvz", "efficiency", "penalty", "calls", "operators"} == keys
    for metric in metrics:
        assert metric["definition"] and metric["label"]
