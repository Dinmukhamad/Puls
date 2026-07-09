"""
Экспорт CSV/XLSX (ТЗ §8): rating, operators, shop-requests, weekly-results,
coin-transactions. Оба формата должны возвращать согласованные данные;
доступ — supervisor (своя группа) / manager / admin, оператору — 403.
"""
from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from tests.conftest import make_operator, make_operator_user
from tests.test_coin_rules_and_group_scope import (
    _login,
    _make_group,
    _make_operator_in_group,
    _make_role_user,
)
from tests.test_weekly_accrual_engine import _reset_coin_rules


def _xlsx_rows(content: bytes) -> list[tuple]:
    wb = load_workbook(BytesIO(content))
    return list(wb.active.iter_rows(values_only=True))


def test_export_rating_csv_and_xlsx_match(client, db_session):
    from app.models import entities as m

    week_start, week_end = date(2026, 8, 3), date(2026, 8, 9)
    _reset_coin_rules(client)
    op = make_operator(db_session, full_name="ЭкспортРейтинг")
    db_session.add(m.WeeklyResult(
        operator_id=op.id, week_start=week_start, week_end=week_end,
        final_score=90, contest_points=90, quality_score=92, lateness_count=0, violation_count=0,
    ))
    db_session.commit()

    r_csv = client.get(f"/api/exports/rating?period_start={week_start}&period_end={week_end}&format=csv")
    assert r_csv.status_code == 200, r_csv.text
    assert r_csv.headers["content-type"].startswith("text/csv")
    assert "ЭкспортРейтинг" in r_csv.text
    assert r_csv.text.startswith("\ufeff")  # BOM — кириллица не ломается (ТЗ 8.6)

    r_xlsx = client.get(f"/api/exports/rating?period_start={week_start}&period_end={week_end}&format=xlsx")
    assert r_xlsx.status_code == 200, r_xlsx.text
    rows = _xlsx_rows(r_xlsx.content)
    assert rows[0][0] == "Место"
    data_row = next(r for r in rows[1:] if "ЭкспортРейтинг" in str(r[1]))
    assert data_row[3] == 90  # Баллы


def test_export_rating_supervisor_scoped_to_group(db_session, make_client):
    from app.models import entities as m

    week_start, week_end = date(2026, 8, 10), date(2026, 8, 16)
    group_a = _make_group(db_session, "ExportRatingA")
    group_b = _make_group(db_session, "ExportRatingB")
    op_a = _make_operator_in_group(db_session, group_a, full_name="РейтингА")
    op_b = _make_operator_in_group(db_session, group_b, full_name="РейтингБ")
    db_session.add_all([
        m.WeeklyResult(operator_id=op_a.id, week_start=week_start, week_end=week_end, final_score=50, contest_points=50),
        m.WeeklyResult(operator_id=op_b.id, week_start=week_start, week_end=week_end, final_score=60, contest_points=60),
    ])
    db_session.commit()

    supervisor, pwd = _make_role_user(db_session, role="supervisor", group_id=group_a.id)
    sup_client = _login(make_client, supervisor.username, pwd)

    r = sup_client.get(f"/api/exports/rating?period_start={week_start}&period_end={week_end}&format=csv")
    assert r.status_code == 200, r.text
    assert "РейтингА" in r.text
    assert "РейтингБ" not in r.text


def test_export_operators_csv(client, db_session):
    make_operator(db_session, full_name="ЭкспортОператоры")
    r = client.get("/api/exports/operators?format=csv")
    assert r.status_code == 200, r.text
    assert "ЭкспортОператоры" in r.text
    assert "ФИО" in r.text


def test_export_shop_requests_status_filter(client, db_session):
    from app.models import entities as m

    op = make_operator(db_session, full_name="ЭкспортЗаявки")
    item = m.ShopItem(title="Экспорт-товар", price=15)
    db_session.add(item)
    db_session.commit()
    db_session.refresh(item)
    p_new = m.ShopPurchase(operator_id=op.id, shop_item_id=item.id, price=15, status="new")
    p_approved = m.ShopPurchase(operator_id=op.id, shop_item_id=item.id, price=15, status="approved")
    db_session.add_all([p_new, p_approved])
    db_session.commit()

    r_all = client.get("/api/exports/shop-requests?format=csv")
    assert r_all.text.count("Экспорт-товар") == 2

    r_new = client.get("/api/exports/shop-requests?status=new&format=csv")
    assert r_new.text.count("Экспорт-товар") == 1
    assert "Новая" in r_new.text


def test_export_weekly_results_before_and_after_apply(client, db_session):
    from app.models import entities as m

    week_start, week_end = date(2026, 8, 17), date(2026, 8, 23)
    _reset_coin_rules(client)
    op = make_operator(db_session, full_name="ЭкспортИтоги")
    db_session.add(m.WeeklyResult(
        operator_id=op.id, week_start=week_start, week_end=week_end,
        final_score=55, contest_points=55, lateness_count=0, violation_count=0,
    ))
    db_session.commit()

    # до apply — предварительный расчёт (тот же движок, что preview)
    r_before = client.get(f"/api/exports/weekly-results?period_start={week_start}&period_end={week_end}&format=csv")
    assert r_before.status_code == 200, r_before.text
    assert "ЭкспортИтоги" in r_before.text

    r_apply = client.post("/api/weekly-results/apply", json={
        "period_start": str(week_start), "period_end": str(week_end), "mode": "manual",
    })
    assert r_apply.status_code == 200, r_apply.text

    r_after = client.get(f"/api/exports/weekly-results?period_start={week_start}&period_end={week_end}&format=csv")
    assert r_after.status_code == 200, r_after.text
    assert "ЭкспортИтоги" in r_after.text


def test_export_coin_transactions_alias_matches_original(client, db_session):
    op = make_operator(db_session, full_name="ЭкспортКоины")
    from app.models import entities as m
    db_session.add(m.CoinTransaction(operator_id=op.id, amount=12, type="manual_add", comment="алиас-тест"))
    db_session.commit()

    r = client.get(f"/api/exports/coin-transactions?operator_id={op.id}&format=csv")
    assert r.status_code == 200, r.text
    assert "ЭкспортКоины" in r.text
    assert "алиас-тест" in r.text


def test_operator_cannot_export_anything(make_client, db_session):
    _, user, pwd = make_operator_user(db_session)
    op_client = _login(make_client, user.username, pwd)

    for url in (
        "/api/exports/rating?period_start=2026-08-01&period_end=2026-08-07",
        "/api/exports/operators",
        "/api/exports/shop-requests",
        "/api/exports/weekly-results?period_start=2026-08-01&period_end=2026-08-07",
        "/api/exports/coin-transactions",
    ):
        r = op_client.get(url)
        assert r.status_code == 403, f"{url} -> {r.status_code}"


def test_invalid_format_rejected(client):
    r = client.get("/api/exports/operators?format=pdf")
    assert r.status_code == 400
